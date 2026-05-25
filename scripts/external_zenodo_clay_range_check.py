from __future__ import annotations

import math
import re
import statistics
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW = ROOT / "external_data" / "zenodo_italian_clays_3600964" / "Italian_Clays_Archive.xlsx"
SOURCE_URL = "https://zenodo.org/records/3600964"


def as_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".")
    if text in {"", "-", " - ", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_model_stiffness() -> tuple[float, float, float]:
    metrics = pd.read_csv(DATA / "external_validation_metrics.csv")
    row = metrics[metrics["source"].str.contains("Zenodo Italian clays", regex=False)].iloc[0]
    values = dict(re.findall(r"(initial|final|loss_pct)=([0-9.]+)", row["model_values"]))
    return float(values["initial"]), float(values["final"]), float(values["loss_pct"])


def stats(values: list[float]) -> dict[str, float | int]:
    clean = sorted(v for v in values if v is not None and math.isfinite(v))
    if not clean:
        return {"n": 0, "min": math.nan, "p05": math.nan, "median": math.nan, "p95": math.nan, "max": math.nan}
    return {
        "n": len(clean),
        "min": min(clean),
        "p05": clean[int(0.05 * (len(clean) - 1))],
        "median": statistics.median(clean),
        "p95": clean[int(0.95 * (len(clean) - 1))],
        "max": max(clean),
    }


def read_dataset_rows() -> list[dict[str, object]]:
    if not RAW.exists():
        raise FileNotFoundError(
            f"Missing {RAW}. Download Italian_Clays_Archive.xlsx from {SOURCE_URL} "
            "into external_data/zenodo_italian_clays_3600964 before running this check."
        )

    wb = load_workbook(RAW, read_only=True, data_only=True)
    ws = wb["Dataset"]
    rows: list[dict[str, object]] = []
    current: dict[str, str] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        values = list(row)
        for idx, key in [(0, "No"), (1, "Site"), (2, "BH"), (6, "Sample"), (7, "Test")]:
            if idx < len(values) and values[idx] not in (None, ""):
                current[key] = str(values[idx]).strip()

        record = {key: current.get(key) for key in ("No", "Site", "BH", "Sample", "Test")}
        record.update(
            {
                "G0c_MPa": as_float(values[23] if len(values) > 23 else None),
                "D0c_pct": as_float(values[24] if len(values) > 24 else None),
                "gamma_RC_pct": as_float(values[25] if len(values) > 25 else None),
                "G_RC_MPa": (as_float(values[26] if len(values) > 26 else None) or math.nan) / 1000.0,
                "D_RC_pct": as_float(values[27] if len(values) > 27 else None),
                "gamma_CTS_pct": as_float(values[29] if len(values) > 29 else None),
                "G_CTS_MPa": (as_float(values[30] if len(values) > 30 else None) or math.nan) / 1000.0,
                "D_CTS_pct": as_float(values[31] if len(values) > 31 else None),
            }
        )
        if any(
            record[key] is not None and not (isinstance(record[key], float) and math.isnan(record[key]))
            for key in ("G0c_MPa", "D0c_pct", "G_RC_MPa", "D_RC_pct", "G_CTS_MPa", "D_CTS_pct")
        ):
            rows.append(record)
    return rows


def main() -> None:
    DATA.mkdir(exist_ok=True)
    rows = read_dataset_rows()
    model_initial, model_final, model_loss = parse_model_stiffness()

    g0 = [r["G0c_MPa"] for r in rows if isinstance(r["G0c_MPa"], float)]
    d0 = [r["D0c_pct"] for r in rows if isinstance(r["D0c_pct"], float)]
    dynamic_g = [
        v
        for r in rows
        for v in (r["G_RC_MPa"], r["G_CTS_MPa"])
        if isinstance(v, float) and math.isfinite(v)
    ]
    dynamic_d = [
        v
        for r in rows
        for v in (r["D_RC_pct"], r["D_CTS_pct"])
        if isinstance(v, float) and math.isfinite(v)
    ]
    strain = [
        v
        for r in rows
        for v in (r["gamma_RC_pct"], r["gamma_CTS_pct"])
        if isinstance(v, float) and math.isfinite(v)
    ]

    metrics = [
        ("small_strain_G0c_MPa", g0),
        ("dynamic_G_RC_CTS_MPa", dynamic_g),
        ("small_strain_D0c_pct", d0),
        ("dynamic_D_RC_CTS_pct", dynamic_d),
        ("strain_RC_CTS_pct", strain),
    ]

    out_rows = []
    for metric, values in metrics:
        applies_to_model_stiffness = metric in {"small_strain_G0c_MPa", "dynamic_G_RC_CTS_MPa"}
        row = {
            "source": "Zenodo Italian clays RC/CTS dataset",
            "source_url": SOURCE_URL,
            "license": "CC BY 4.0",
            "raw_file": str(RAW.relative_to(ROOT)).replace("\\", "/"),
            "rows_with_data": len(rows),
            "unique_samples_with_dynamic_rows": len({r["Sample"] for r in rows if r.get("Sample")}),
            "metric": metric,
            **stats(values),
            "model_initial_mpa": model_initial,
            "model_final_mpa": model_final,
            "model_stiffness_loss_pct": model_loss,
            "model_initial_inside_range": min(values) <= model_initial <= max(values) if values and applies_to_model_stiffness else "",
            "model_final_inside_range": min(values) <= model_final <= max(values) if values and applies_to_model_stiffness else "",
            "comparison_scope": "external dynamic-stiffness range check only; not calibration or blind validation",
        }
        out_rows.append(row)

    out = pd.DataFrame(out_rows)
    out.to_csv(DATA / "zenodo_italian_clays_range_check.csv", index=False)
    print(f"zenodo_italian_clays_range_check=ok rows={len(out)} raw_rows={len(rows)}")


if __name__ == "__main__":
    main()
